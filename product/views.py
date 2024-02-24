from advertisement.models import Advertisement
from advertisement.serializers import AdvertisementSerializer
from common.utils.create_slug import create_slug
from django.core.exceptions import ObjectDoesNotExist
from django.db import IntegrityError, transaction
from django.db.models import Case, ExpressionWrapper, F, FloatField, Sum, When
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import load_workbook
from rest_framework import filters, status, viewsets
from rest_framework.generics import (
    DestroyAPIView,
    GenericAPIView,
    ListAPIView,
    RetrieveUpdateDestroyAPIView,
)
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .filters import ProductFilter
from .mixins import CheckProductManagerGroupMixin, CheckSupplierAdminGroupMixin
from .models import Brand, Category, PrivateCategory, Product
from .serializers import (
    BrandSerializer,
    CategorySerializer,
    PrivateCategorySerializer,
    ProductSerializer,
)


class CategoryViewSet(CheckProductManagerGroupMixin, viewsets.ViewSet):
    queryset = Category.objects.all()

    def list(self, request):
        level = request.GET.get("level")
        featured = request.GET.get("featured")

        if featured == "true":
            serializer = CategorySerializer(self.queryset.filter(is_featured=True), many=True)

        else:
            serializer = CategorySerializer(self.queryset.all(), many=True)

            if request.GET.get("parent"):
                serializer = CategorySerializer(
                    self.queryset.filter(level=level, parent__slug=request.GET.get("parent")),
                    many=True,
                )
            elif level:
                serializer = CategorySerializer(
                    self.queryset.filter(level=level),
                    many=True,
                )

        return Response(serializer.data, status=status.HTTP_200_OK)


class BrandViewSet(CheckProductManagerGroupMixin, viewsets.ModelViewSet):
    queryset = Brand.objects.all()
    serializer_class = BrandSerializer

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        limit = int(request.GET.get("limit")) if request.GET.get("limit") else None

        if limit:
            queryset = queryset[:limit]

        serializer = self.get_serializer(queryset, many=True, context={"request": request})
        return Response(serializer.data)


class ProductViewSet(CheckProductManagerGroupMixin, viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = ProductFilter
    ordering_fields = [
        "created",
        "name",
        "blended_price",
    ]
    ordering = "name"

    def get_queryset(self):
        queryset = super().get_queryset()

        if self.request.user.is_anonymous or self.request.user.is_buyer:
            queryset = queryset.filter(is_available=True, stock_quantity__gt=0)

        ordering_param = self.request.query_params.get("order")

        if ordering_param == "best_selling":
            queryset = queryset.annotate(total_sold=Sum("orderitem__quantity")).order_by(
                "-total_sold"
            )
        else:
            queryset = queryset.annotate(
                blended_price=ExpressionWrapper(
                    Case(
                        When(price=0, then=F("price")),
                        When(sale_price=0, then=F("sale_price")),
                        When(
                            price_range_min=0,
                            price_range_max=0,
                            then=(F("price_range_min") + F("price_range_max")) / 2,
                        ),
                        output_field=FloatField(),
                    ),
                    output_field=FloatField(),
                )
            )

        return queryset

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        if request.query_params.get("ads", "").lower() == "true":
            ads_queryset = Advertisement.objects.all()
            ads_serializer = AdvertisementSerializer(
                ads_queryset, many=True, context={"request": request}
            )

            serializer = self.get_serializer(queryset, many=True, context={"request": request})

            data = {
                "products": serializer.data,
                "ads": ads_serializer.data,
            }

            return Response(data)

        serializer = self.get_serializer(queryset, many=True, context={"request": request})
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        try:
            data = request.data

            category_name = data.pop("category[label]")
            sub_category_name = data.pop("subCategory[label]")
            brand_name = data.pop("brand[label]")

            category = Category.objects.filter(name__iexact=category_name[0]).first()

            if not category:
                category = Category.objects.create(
                    name=category_name[0], image=data.pop("categoryImage")[0]
                )

            sub_category = Category.objects.filter(
                name__iexact=sub_category_name[0], parent=category
            ).first()

            if not sub_category:
                sub_category = Category.objects.create(
                    name=sub_category_name[0],
                    parent=category,
                    image=data.pop("subCategoryImage")[0],
                )

            brand = Brand.objects.filter(name__iexact=brand_name[0]).first()

            if not brand:
                brand = Brand.objects.create(
                    name=brand_name[0], image=data.pop("brandImage")[0]
                )

            thumbnail = data.pop("newThumbnail", None)
            other_images = data.getlist("otherImages[]", None)

            product_data = {
                "supplier": request.user,
                "name": data["name"],
                "description": data["description"],
                "price": data["price"],
                "sale_price": data["salePrice"],
                "price_range_min": data["priceRangeMin"],
                "price_range_max": data["priceRangeMax"],
                "category": sub_category,
                "brand": brand,
                "stock_quantity": data["stockQuantity"],
                "is_available": True if data["isAvailable"] == "true" else False,
                "thumbnail": thumbnail[0],
                "image1": other_images[0] if len(other_images) >= 1 else None,
                "image2": other_images[1] if len(other_images) >= 2 else None,
                "image3": other_images[2] if len(other_images) >= 3 else None,
                "image4": other_images[3] if len(other_images) >= 4 else None,
                "is_returnable": True if data["isReturnable"] == "true" else False,
                "return_deadline": data["returnDeadline"] if data["returnDeadline"] else 30,
            }

            Product.objects.create(**product_data)

            return Response(status=status.HTTP_201_CREATED)

        except Exception as ex:
            print(ex)
            return Response(status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        try:
            data = request.data.copy()

            category_name = data.pop("category[label]", None)[0]
            sub_category_name = data.pop("subCategory[label]", None)[0]
            brand_name = data.pop("brand[label]", None)[0]

            category = Category.objects.filter(name__iexact=category_name).first()

            if not category:
                category = Category.objects.create(
                    name=category_name, image=data.pop("categoryImage")[0]
                )

            sub_category = Category.objects.filter(
                name__iexact=sub_category_name, parent=category
            ).first()

            if not sub_category:
                sub_category = Category.objects.create(
                    name=sub_category_name,
                    parent=category,
                    image=data.pop("subCategoryImage")[0],
                )

            brand = Brand.objects.filter(name__iexact=brand_name).first()

            if not brand:
                brand = Brand.objects.create(name=brand_name, image=data.pop("brandImage")[0])

            thumbnail = data.pop("newThumbnail", None)
            newImage1 = data.pop("newImage1", None)
            newImage2 = data.pop("newImage2", None)
            newImage3 = data.pop("newImage3", None)
            newImage4 = data.pop("newImage4", None)

            product_data = {
                "name": data.get("name"),
                "description": data.get("description"),
                "price": data.get("price"),
                "sale_price": data.get("salePrice"),
                "price_range_min": data.get("priceRangeMin"),
                "price_range_max": data.get("priceRangeMax"),
                "category": sub_category,
                "brand": brand,
                "stock_quantity": data.get("stockQuantity"),
                "is_available": True if data.get("isAvailable") == "true" else False,
                "is_returnable": True if data.get("isReturnable") == "true" else False,
                "return_deadline": (
                    data.get("returnDeadline") if data.get("returnDeadline") else 30
                ),
            }

            product = Product.objects.get(sku=kwargs["pk"])

            for key, value in product_data.items():
                if value is not None:
                    setattr(product, key, value)

            if thumbnail:
                product.thumbnail = thumbnail[0]

            if newImage1:
                product.image1 = newImage1[0]

            if newImage2:
                product.image2 = newImage2[0]

            if newImage3:
                product.image3 = newImage3[0]

            if newImage4:
                product.image4 = newImage4[0]

            product.save()

            return Response(status=status.HTTP_200_OK)

        except Exception as ex:
            print(ex)
            return Response(status=status.HTTP_400_BAD_REQUEST)


class DeleteProductView(CheckProductManagerGroupMixin, DestroyAPIView):
    def destroy(self, request, *args, **kwargs):
        data = request.data

        if str(data["user"]) == str(request.user.id):
            try:
                product = Product.objects.get(sku=str(data["sku"]))

                product.delete()

                return Response(status=status.HTTP_204_NO_CONTENT)
            except Exception as ex:
                print(ex)
                pass

        return Response(status=status.HTTP_400_BAD_REQUEST)


class GetSupplierProductsView(CheckProductManagerGroupMixin, ListAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer

    def get_queryset(self):
        return (
            super().get_queryset().filter(supplier__id__iexact=self.kwargs.get("id")).distinct()
        )


class CreateProductExcelView(CheckProductManagerGroupMixin, GenericAPIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES["excelFile"]

        wb = load_workbook(excel_file)
        ws = wb.active

        with transaction.atomic():
            for row in ws.iter_rows(min_row=2, values_only=True):
                (
                    name,
                    description,
                    price,
                    sale_price,
                    price_range_min,
                    price_range_max,
                    category_name,
                    subcategory_name,
                    brand_name,
                    stock_quantity,
                    return_deadline,
                ) = row

                if stock_quantity == 0:
                    is_available = False
                else:
                    is_available = True

                if return_deadline == 0:
                    is_returnable = False
                else:
                    is_returnable = True

                thumbnail = "/product/images/placeholder/placeholder.png"

                try:
                    category = Category.objects.get(slug=create_slug(category_name))
                except Category.DoesNotExist:
                    try:
                        category = Category.objects.create(
                            name=category_name,
                            slug=create_slug(category_name),
                        )
                    except IntegrityError:
                        category = Category.objects.get(slug=create_slug(category_name))

                try:
                    subcategory = Category.objects.get(slug=create_slug(subcategory_name))
                except Category.DoesNotExist:
                    try:
                        subcategory = Category.objects.create(
                            name=subcategory_name,
                            parent=category,
                            slug=create_slug(subcategory_name),
                        )
                    except IntegrityError:
                        subcategory = Category.objects.get(slug=create_slug(subcategory_name))

                try:
                    brand = Brand.objects.get(slug=create_slug(brand_name))
                except Brand.DoesNotExist:
                    try:
                        brand = Brand.objects.create(
                            name=brand_name,
                            slug=create_slug(brand_name),
                        )
                    except IntegrityError:
                        brand = Brand.objects.get(slug=create_slug(brand_name))

                if price > 0:
                    if price_range_min != 0 or price_range_max != 0:
                        return Response(
                            {
                                "error": "Both minimum and maximum price range should be 0 when price is greater than 0."
                            },
                            status=status.HTTP_400_BAD_REQUEST,
                        )

                    if price_range_min == 0 and price_range_max == 0:
                        return Response(
                            {
                                "error": "Either minimum or maximum price range should be greater than 0 when price is greater than 0."
                            },
                            status=status.HTTP_400_BAD_REQUEST,
                        )

                    if (
                        price_range_min <= 0
                        or price_range_max <= 0
                        or price_range_min > price_range_max
                    ):
                        return Response(
                            {"error": "Invalid minimum and maximum price range values."},
                            status=status.HTTP_400_BAD_REQUEST,
                        )

                try:
                    Product.objects.create(
                        name=name,
                        description=description,
                        supplier=request.user,
                        price=price,
                        sale_price=sale_price,
                        price_range_min=price_range_min,
                        price_range_max=price_range_max,
                        category=subcategory,
                        brand=brand,
                        stock_quantity=stock_quantity,
                        is_available=is_available,
                        return_deadline=return_deadline,
                        is_returnable=is_returnable,
                        thumbnail=thumbnail,
                    )

                except Exception as ex:
                    print(ex)
                    return Response(status=status.HTTP_400_BAD_REQUEST)

            return Response(status=status.HTTP_201_CREATED)


class PrivateCategoryAPIView(CheckSupplierAdminGroupMixin, GenericAPIView):
    queryset = PrivateCategory.objects.all()
    serializer_class = PrivateCategorySerializer

    def get_queryset(self):
        user = self.request.user
        return self.queryset.filter(supplier=user)

    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        category_serializer = self.get_serializer(
            queryset, many=True, context={"request": request}
        )

        products = Product.objects.filter(supplier=request.user)
        product_serializer = ProductSerializer(
            products, many=True, context={"request": request}
        )

        category_data = category_serializer.data
        product_data = product_serializer.data

        response_data = {
            "private_categories": category_data,
            "products": product_data,
        }

        return Response(response_data, status=status.HTTP_200_OK)

    def post(self, request, *args, **kwargs):
        if request.data.get("categoryName"):
            try:
                PrivateCategory.objects.create(
                    name=request.data.get("categoryName"), supplier=request.user
                )
                return Response(status=status.HTTP_201_CREATED)
            except Exception as ex:
                return Response(str(ex), status=status.HTTP_400_BAD_REQUEST)

        elif request.data.get("product_id"):
            if request.data.get("add") == "true":
                try:
                    product_id = request.data.get("product_id")
                    product = Product.objects.get(sku=product_id)

                    category_id = request.data.get("category_id")
                    category = PrivateCategory.objects.get(id=category_id)

                    category.products.add(product)

                    category.save()

                    return Response(status=status.HTTP_201_CREATED)

                except Exception as ex:
                    print(ex)
                    return Response(str(ex), status=status.HTTP_400_BAD_REQUEST)

            else:
                try:
                    product_id = request.data.get("product_id")
                    product = Product.objects.get(sku=product_id)

                    category_id = request.data.get("category_id")
                    category = PrivateCategory.objects.get(id=category_id)

                    category.products.remove(product)

                    category.save()

                    return Response(status=status.HTTP_201_CREATED)

                except Exception as ex:
                    print(ex)
                    return Response(str(ex), status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, *args, **kwargs):
        private_category_id = request.data.get("categoryID")

        try:
            category = PrivateCategory.objects.get(id=private_category_id)
            category.delete()
            return Response(status=status.HTTP_200_OK)
        except ObjectDoesNotExist:
            return Response({"error": "Category not found"}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=500)


class PrivateCategoryDetailAPIView(CheckSupplierAdminGroupMixin, RetrieveUpdateDestroyAPIView):
    queryset = PrivateCategory.objects.all()
    serializer_class = PrivateCategorySerializer
